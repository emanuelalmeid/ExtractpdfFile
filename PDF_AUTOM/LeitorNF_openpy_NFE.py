# importa as bibliotecas necessárias
import openpyxl
from pdfminer.high_level import extract_text
import re
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import matplotlib as plt 
#Variáveis
i = 2 #primeira célula (2 caso seja do ínicio)
UltimaCell = 120 #Ultima célula a planilha excel
# Abre o arquivo pdf 
# lembre-se que para o windows você deve usar essa barra -> / 
# lembre-se também que você precisa colocar o caminho absoluto


wb = load_workbook('C:\\Users\\User\\Python_projects\\LeitorNF\\BASENOTASNFE.xlsx')
ws = wb['BASENOTAS']
for i in range(i, UltimaCell+1):
    NFE = ws["A"+ str(i)].value
    texto = extract_text("C:\\Users\\User\\Python_projects\\NOTAS\\"+ str(NFE) + "NFE.pdf")
#adicionar if caso retorne algo
    index_dadosADIC = texto.find("TRANSPORTADOR/VOLUME TRANSPORTADOS")
    index_fim = texto.find("MOTORISTA") 
    dadosADIC = texto[index_dadosADIC: index_fim]
    index_dadosADIC = texto.find("NOME/RAZÃO")
    dadosADIC = texto[index_dadosADIC: index_fim]

    index_dadosADIC = texto.find("INFORMAÇÕES COMPLEMENTARES")
    index_fim = texto.find("RESERVADO AO FISCO") 
    dadosADIC2 = texto[index_dadosADIC: index_fim]

    index_dadosADIC = texto.find("NOME FANTASIA")
    index_fim = texto.find("CÓDIGO DO CLIENTE") 
    dadosADIC3 = texto[index_dadosADIC+13: index_fim]


    index_dadosADIC = texto.find("Mapa:")
    index_fim = texto.find("DANFE") 
    dadosADIC4 = texto[index_dadosADIC: index_fim]





    dadosADIC = re.sub('/n','',dadosADIC)
    dadosADIC2 = re.sub('/n','',dadosADIC2)

    ws["E"+ str(i)] = dadosADIC4
    ws["D"+ str(i)] = dadosADIC3
    ws["C"+ str(i)] = dadosADIC2
    ws["B"+ str(i)] = dadosADIC
    
    print(i)
    #print(ws["C"+ str(i)].value)



print(dadosADIC3)
wb.save('C:\\Users\\User\\Python_projects\\LeitorNF\\BASENOTASNFE.xlsx')
