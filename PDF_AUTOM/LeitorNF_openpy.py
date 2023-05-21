# importa as bibliotecas necessárias
import openpyxl
from pdfminer.high_level import extract_text
import re
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter

#Variáveis
i = 2
UltimaCell = 248 #Ultima célula a planilha excel
# Abre o arquivo pdf 
# lembre-se que para o windows você deve usar essa barra -> / 
# lembre-se também que você precisa colocar o caminho absoluto


wb = load_workbook('C:\\Users\\User\\Python_projects\\LeitorNF\\BASENOTAS.xlsx')
ws = wb['BASENOTAS']
for i in range(i, UltimaCell+1):
    NFD = ws["A"+ str(i)].value
    texto = extract_text("C:\\Users\\User\\Python_projects\\NOTAS\\"+ str(NFD) + "NFD.pdf")
#adicionar if caso retorne algo
    index_dadosADIC = texto.find("INFORMAÇÕES COMPLEMENTARES")
    index_fim = texto.find("RESERVADO AO FISCO") 
    dadosADIC = texto[index_dadosADIC: index_fim]



    index_dadosADIC = texto.find("Mapa:")
    index_fim = texto.find("DANFE") 
    dadosADIC3 = texto[index_dadosADIC: index_fim]
    dadosADIC = re.sub('/n','',dadosADIC)





    ws["B"+ str(i)] = dadosADIC
    ws["E"+ str(i)] = dadosADIC3
    #print(ws["B"+ str(i)].value)
    print(i)



print("a")
wb.save('C:\\Users\\User\\Python_projects\\LeitorNF\\BASENOTAS.xlsx')
